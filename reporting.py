import os.path
import openpyxl


def create_report(site_name, pg_names):
    wb = openpyxl.Workbook()
    wb.create_sheet(f'PG Report', -1)
    headers = [i for i in pg_names]
    headers.insert(0, 'Device')
    ws = wb['PG Report']

    for i in range(len(headers)):
        ws.cell(row=1, column=i+1, value=headers[i])
    directories = os.listdir()
    for directory in directories:
        if directory == site_name and os.path.isdir(directory):
            sub_dirs = os.listdir(directory)
            device_list = []
            for sub_dir in sub_dirs:
                path = f'{directory}/{sub_dir}'
                if os.path.isdir(path):
                    type_dirs = os.listdir(path)
                    for type_dir in type_dirs:
                        path = f'{directory}/{sub_dir}/{type_dir}'
                        if os.path.isdir(path):
                            files = os.listdir(path)
                            for file in range(len(files)):
                                file_name_parts = files[file].split(sep=' - ')
                                device, pg_inst, pg_file_name = file_name_parts[0], file_name_parts[1], file_name_parts[2].split(sep='.')
                                for header in headers:
                                    if pg_file_name[0] == header:
                                        if device in device_list:
                                            continue
                                        else:
                                            device_list.append({device: {pg_file_name[0]: type_dir}})
    row_count = 2
    devs = [key for item in device_list for key, value in item.items()]
    devs = list(dict.fromkeys(devs))
    for dev in sorted(devs):
        ws.cell(row=row_count, column=1, value=dev)
        row_count += 1
    for row in range(len(devs)):
        cell_value = ws.cell(row=row+2, column=1).value
        for item in device_list:
            for key, value in item.items():
                for k, v in value.items():
                    if cell_value == key:
                        for h, i in enumerate(headers):
                            if i == k:
                                ws.cell(row=row+2, column=h+1, value=v)
    wb.save(f'PG Report - {site_name}.xlsx')


# def test_walk():
#     for path, directories, files in os.walk('.\\OSU Something'):
#         print(f'Path: {path}\nDirectories: {directories}\nFile Names: {files}\n\n')
#
#
# test_walk()
